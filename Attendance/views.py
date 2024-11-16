from datetime import timedelta
from django.views.decorators.csrf import csrf_protect
import openpyxl 
from django.http import HttpResponse
import numpy as np
import io
from datetime import date
from django.utils import timezone
from django.shortcuts import redirect, render
from django.contrib.auth import authenticate,login,logout
from .models import Students,Attendance
from django.db import transaction
from .utils import capture_face_data,compare_faces


# Create your views here.
def home(request):
    today = timezone.now().date()
    attendance_exists = Attendance.objects.filter(date=today).exists()
    if not attendance_exists:
        try:
            with transaction.atomic():
                students = Students.objects.all()
                attendance_records = [Attendance(student=student,date=today,)for student in students]
                Attendance.objects.bulk_create(attendance_records)  
                print("done")
        except Exception as e:
            print(e)    
            
    # DASHBOARD
    student_count=Students.objects.all().count()
    attendance=Attendance.objects.filter(date=date.today()).count()
    present=Attendance.objects.filter(date=date.today(), remark="Present").count()
    if attendance>0:
        attendence_per=(present/attendance)*100
    else:
        attendence_per=0
    absent=student_count-present
    context={"student_count":student_count,"attendence_per":attendence_per,"absent":absent}
    
          
    return render(request,'home.html',context)

def logoutUser(request):
    logout(request)
    return redirect("Attendance:home")

@csrf_protect
def Report(request):
    if request.method=="POST":
        students = Students.objects.all()
    
        # Get all attendance records
        attendance_records = Attendance.objects.all()
        
        # Create a new workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance"
        
        # Add row headers (student names)
        for row, student in enumerate(students, start=2):
            worksheet.cell(row=row, column=1, value=str(student.name))
        
        # Add column headers (dates)
        dates = sorted(set(attendance.date for attendance in attendance_records))
        for col, date in enumerate(dates, start=2):
            worksheet.cell(row=1, column=col, value=date)
        
        # Populate the worksheet with data
        for row, student in enumerate(students, start=2):
            for col, date in enumerate(dates, start=2):
                try:
                    record = Attendance.objects.get(date=date, student=student)
                    worksheet.cell(row=row, column=col, value=record.remark)
                except Attendance.DoesNotExist:
                    worksheet.cell(row=row, column=col, value="")
        
        # Create the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
        workbook.save(response)
        
        return response
    else:
    
        return render(request,'Report.html')


def loginuser(request):
    if request.user.is_authenticated:
        return redirect("Attendance:home")  
        
    if request.method=="POST":
        username=request.POST.get('username')
        password=request.POST.get('password')   
        a=authenticate(request,username=username,password=password) 
        if a is not None:
            login(request,a)  
            return redirect("Attendance:manage")
        else:
            message={"message":"You are not authorized"}
            return render(request,"login.html",message)
    return render(request,"login.html")



def manageStudents(request):
    if request.user.is_authenticated:
        time_diff = timezone.now() - request.user.last_login
        if time_diff > timedelta(hours=2):
            logout(request)
            return render(request,"login.html")
        #ADDING STUDENT
        if request.method=="POST":
            face_data=capture_face_data()
            path=f"C:\\Users\\91983\\Desktop\\Django\\FaceR\\Attendance\\static\\Students\\{request.POST.get("user_id")}"
            if face_data is None:
                return render(request,"manage.html",{"message":"Some Error Occured Try Again"})
            Students.objects.create(
            name=request.POST.get("name"),
            user_id=request.POST.get("user_id"),
            password=request.POST.get("password"),
            email=request.POST.get("email"),
            phone_number=request.POST.get("phone_number"))
            with open(path,"w") as file:
                file.write(str(face_data))
            return render(request,"manage.html")
            
        return render(request,"manage.html")
    else:
        return redirect("Attendance:Login")

def loginStudent(request):
    if request.user.is_authenticated:
        return redirect("Attendance:home")  
        
    if request.method=="POST":
        user_id=request.POST.get('username')
        password=request.POST.get('password')   
        student=Students.objects.get(user_id=user_id)
        if student is not None:
            if student.password==password:
                request.session['user_id'] = user_id
                request.session['name'] = student.name
                return redirect('Attendance:MarkAttendance')
            else:   
                message={"message":"Invalid Credentials"}
                return render(request,"Studentlogin.html",message) 
        else:
            message={"message":"You are not authorized"}
            return render(request,"Studentlogin.html",message)
    return render(request,"Studentlogin.html")

def markattendance(request):
    user_id = request.session.get('user_id')
    name = request.session.get('name')
    student=Students.objects.get(user_id=user_id)
    if request.method=="POST":
        try:
            path=f"Attendance\\static\\Students\\{user_id}"
            data1=capture_face_data()
            with open(path,"r")as file:
                array_data1=file.read() 
            array_data1 = array_data1.replace('[', '').replace(']', '').strip() 
            data2 = np.fromstring(array_data1, sep=' ')
            if compare_faces(data1,data2) == True:
                print("done")
                attendance=Attendance.objects.filter(date=date.today(),student=student)
                attendance.update(remark="Present")
                context = {"user_id": user_id, "name": name,"message":"Success"}
                print("done2")
                return render(request,"MarkAttendance.html",context)
            else:
                print("not done")
                context = {"user_id": user_id, "name": name,"message":"Fail"}
                return render(request,"MarkAttendance.html",context)
        except Exception as e :
            print("Failed")
            context = {"user_id": user_id, "name": name,"message":"Fail"}
            return render(request,"MarkAttendance.html",context)
    else:        
    
        context = {"user_id": user_id, "name": name}
        return render(request,"MarkAttendance.html",context)

    
    